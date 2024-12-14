import openpyxl
from tkinter import Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename

def delete_and_shift_rows(file_path, strings_to_delete, output_file_path, log_file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)
    deleted_strings = []

    # Iterate through all sheets
    for sheet in workbook.worksheets:
        rows_to_check = sheet.max_row

        # Traverse the rows from top to bottom
        for row_idx in range(1, rows_to_check + 1):
            row_deleted = False  # To track if the row needs adjustment
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                # Check if the cell value contains any of the strings to delete
                if isinstance(cell.value, str):  # Ensure cell value is a string
                    for string in strings_to_delete:
                        if string in cell.value:
                            deleted_strings.append((sheet.title, cell.coordinate, cell.value))  # Record deleted strings
                            row_deleted = True
                            break  # Stop checking once a match is found

            # If any cell in the row matches, delete the row and shift up
            if row_deleted:
                sheet.delete_rows(row_idx)
                rows_to_check -= 1  # Update the row count
                row_idx -= 1  # Adjust the row index to account for the shift

    # Save the updated workbook to the output file
    workbook.save(output_file_path)

    # Write the log of deleted strings
    with open(log_file_path, 'w') as log_file:
        for sheet, cell, value in deleted_strings:
            log_file.write(f"Deleted '{value}' from {sheet} at {cell}\n")

    print(f"Processing complete. Updated file saved to '{output_file_path}'. Log saved to '{log_file_path}'.")

# Main function to handle file dialogs and process
def main():
    # Hide the root tkinter window
    root = Tk()
    root.withdraw()

    # Open file dialog to select the input Excel file
    print("Select the Excel file to process:")
    input_file_path = askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not input_file_path:
        print("No file selected. Exiting.")
        return

    # Open file dialog to specify where to save the updated Excel file
    print("Select where to save the updated Excel file:")
    output_file_path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not output_file_path:
        print("No output file selected. Exiting.")
        return

    # Define log file path (you can also ask for a dialog for this if needed)
    log_file_path = output_file_path.replace(".xlsx", "_log.txt")

    # Ask the user how to provide strings to delete
    print("Would you like to load strings to delete from a file? (yes/no)")
    choice = input().strip().lower()

    if choice == "yes":
        print("Select the text file containing strings to delete (one string per line):")
        strings_file_path = askopenfilename(filetypes=[("Text files", "*.txt")])
        if not strings_file_path:
            print("No file selected. Exiting.")
            return

        # Read strings from the file
        with open(strings_file_path, 'r') as f:
            strings_to_delete = [line.strip() for line in f if line.strip()]
    else:
        # Allow manual input
        print("Enter strings to delete, separated by commas:")
        strings_to_delete = [s.strip() for s in input().split(',') if s.strip()]

    # Call the function to process the file
    delete_and_shift_rows(input_file_path, strings_to_delete, output_file_path, log_file_path)

if __name__ == "__main__":
    main()
